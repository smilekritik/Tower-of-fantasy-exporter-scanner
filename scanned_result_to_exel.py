import json
import colorsys
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

def clean_file_name(filename):
    return filename.replace(".json", "")


def generate_rainbow_colors(n):
    """Generate 'n' distinct rainbow colors."""
    colors = []
    for i in range(n):
        hue = i / n
        rgb = colorsys.hsv_to_rgb(hue, 1.0, 1.0)
        hex_color = ''.join(f'{int(c * 255):02X}' for c in rgb)
        colors.append(hex_color)
    return colors


def darken_color(hex_color, factor=2):
    """Darken a color by dividing the RGB values."""
    hex_color = hex_color.lstrip('#')
    rgb = tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    darkened_rgb = tuple(max(int(c // factor), 0) for c in rgb)
    return f'{darkened_rgb[0]:02X}{darkened_rgb[1]:02X}{darkened_rgb[2]:02X}'

def get_category_color(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


attribute_colors = {
    'Final': 'D3D3D3',  # Light gray
    'Phy': 'FFA07A',  # Light orange
    'Fire': 'EC5353',  # Light red
    'Ice': 'ADD8E6',  # Light blue
    'Thunder': 'DDA0DD',  # Light purple
    'Superpower': '90EE90',  # Light green
    'Crit': 'FFFF00',  # Yellow for Crit
    'Common': 'F5F5DC'  # Light beige for Common
}


def get_attribute_fill_color(attribute_name):
    # Priority 1: Crit (must always be yellow if "Crit" is found anywhere)
    if 'Crit' in attribute_name:
        return PatternFill(start_color=attribute_colors['Crit'], end_color=attribute_colors['Crit'], fill_type='solid')

    # Priority 2: Other attribute-specific colors (Final, Phy, Fire, etc.)
    base_color = None
    for key in attribute_colors:
        if attribute_name.startswith(key):
            base_color = attribute_colors[key]
            break

    # If no base color is found, return None
    if not base_color:
        return None

    # Priority 3: Darken the color if Down or Def is found, but Ignore and Break should prevent darkening
    if (
            'Down' in attribute_name or 'Def' in attribute_name) and 'Ignore' not in attribute_name and 'Break' not in attribute_name:
        base_color = darken_color(base_color)  # Darken the color

    return PatternFill(start_color=base_color, end_color=base_color, fill_type='solid')

with open('module_extra_to_files_mapping3.json', 'r', encoding='utf-8') as f:
    json_data = json.load(f)

categories = list(json_data.keys())
rainbow_colors = generate_rainbow_colors(len(categories))


wb = Workbook()
ws = wb.active
ws.title = "ALL_trash"
fixed_width = 60


start_row = 1
start_col = 1


col = start_col
for i, (module_type, attributes) in enumerate(json_data.items()):
    category_color = rainbow_colors[i]

    # Write the ModuleExtraType as a merged header across all attribute columns
    attr_col = col
    file_row = start_row + 1  # File names go one row below attributes

    # Write ModuleExtraType in the first row and merge it across all the columns of its attributes
    end_col = attr_col + len(attributes) - 1
    ws.merge_cells(start_row=start_row, start_column=attr_col, end_row=start_row, end_column=end_col)
    ws.cell(row=start_row, column=attr_col).value = module_type
    ws.cell(row=start_row, column=attr_col).alignment = Alignment(horizontal='center', vertical='center')

    # Apply category coloring (rainbow effect for each category)
    for col_idx in range(attr_col, end_col + 1):
        ws.cell(row=start_row, column=col_idx).fill = get_category_color(category_color)

    # Write all attributes in the same row
    attr_row = start_row + 1
    max_len = 0  # To track how many rows the longest attribute has

    for attribute, file_list in attributes.items():
        # Write attribute in the current column
        ws.cell(row=attr_row, column=attr_col).value = attribute
        ws.cell(row=attr_row, column=attr_col).alignment = Alignment(horizontal='center', vertical='center')

        # Apply conditional color fill to the attribute based on its prefix (color the entire column)
        fill_color = get_attribute_fill_color(attribute)
        if fill_color:
            ws.cell(row=attr_row, column=attr_col).fill = fill_color

        # Write corresponding files vertically under each attribute
        row = attr_row + 1
        for file in file_list:
            clean_file = clean_file_name(file)
            ws.cell(row=row, column=attr_col).value = clean_file
            ws.cell(row=row, column=attr_col).alignment = Alignment(horizontal='center')

            # Apply color to the entire column for the attribute
            if fill_color:
                ws.cell(row=row, column=attr_col).fill = fill_color

            row += 1

        max_len = max(max_len, len(file_list))
        ws.column_dimensions[get_column_letter(attr_col)].width = fixed_width
        attr_col += 1

    col = end_col + 2

output_filename = 'exported_v1.xlsx'
wb.save(output_filename)

print(f"Data successfully written to {output_filename}")